import asyncio
import os
import uuid
from datetime import date, datetime, timedelta
import httpx
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from core.config import settings
from core.wisphub.client import wisphub_client
from modules.finanzas.models import VerificacionPago, LogRecordatorio, PagoRegistrado, RecoleccionRegistro, Observacion  # noqa: F401

GRAPH_URL = "https://graph.facebook.com/v19.0"


def _week_range(fecha_inicio: str | None = None) -> tuple[date, date]:
    if fecha_inicio:
        base = date.fromisoformat(fecha_inicio)
        start = base - timedelta(days=base.weekday())  # Monday of that week
    else:
        today = date.today()
        start = today - timedelta(days=today.weekday())  # Monday
    end = start + timedelta(days=6)  # Sunday
    return start, end

def _tipo_cobro(articulos: list) -> str:
    for art in articulos:
        desc = (art.get("descripcion") or "").lower()
        if "instalaci" in desc:
            return "instalacion"
    return "mensualidad"

async def get_cobros_semana(fecha_inicio: str | None = None) -> dict:
    start, end = _week_range(fecha_inicio)
    start_str = start.isoformat()
    end_str = end.isoformat()

    data = await wisphub_client.get("/api/facturas/", params={"page_size": 1000})

    items = []
    total_monto = 0.0
    total_pagado = 0.0
    total_pendiente = 0.0

    for f in data.get("results", []):
        emision = (f.get("fecha_emision") or "")[:10]
        if not (start_str <= emision <= end_str):
            continue

        monto = float(f.get("total") or 0)
        estado = f.get("estado", "")
        articulos = f.get("articulos", [])
        cliente = f.get("cliente") or {}

        total_monto += monto
        if estado == "Pagada":
            total_pagado += monto
        else:
            total_pendiente += monto

        id_servicio = next(
            (art.get("servicio", {}).get("id_servicio") for art in articulos if art.get("servicio")),
            None
        )
        items.append({
            "id_factura": f["id_factura"],
            "fecha_emision": emision,
            "fecha_vencimiento": (f.get("fecha_vencimiento") or "")[:10],
            "estado": estado,
            "total": monto,
            "tipo_cobro": _tipo_cobro(articulos),
            "cliente": {
                "id_servicio": id_servicio,
                "nombre": cliente.get("nombre", "—"),
                "telefono": cliente.get("telefono", "—"),
                "direccion": cliente.get("direccion", "—"),
            },
        })

    items.sort(key=lambda x: (0 if x["estado"] == "Pendiente de Pago" else 1, x["fecha_vencimiento"]))

    return {
        "semana_inicio": start_str,
        "semana_fin": end_str,
        "count": len(items),
        "total_monto": round(total_monto, 2),
        "total_pagado": round(total_pagado, 2),
        "total_pendiente": round(total_pendiente, 2),
        "items": items,
    }

_ID_A_METODO: dict[int, str] = {
    82219: "efectivo",
    84091: "tarjeta",
    82222: "transferencia bancaria",}

def _metodo_pago(f: dict) -> str:
    raw = f.get("metodo_pago") or f.get("forma_pago") or ""
    if isinstance(raw, dict):
        fid = raw.get("Id") or raw.get("id")
        if fid:
            mapped = _ID_A_METODO.get(int(fid))
            if mapped:
                return mapped
        return (raw.get("Nombre") or raw.get("nombre") or "no_especificado").strip().lower()
    value = str(raw).strip().lower()
    return value if value else "no_especificado"

async def get_cobros_dia(fecha: str | None = None, fecha_fin: str | None = None, db: AsyncSession | None = None) -> dict:
    fecha_str = fecha or date.today().isoformat()
    fecha_fin_str = fecha_fin or fecha_str

    data = await wisphub_client.get("/api/facturas/", params={"page_size": 1000})

    verificaciones: dict[int, bool] = {}
    if db is not None:
        result = await db.execute(select(VerificacionPago))
        for v in result.scalars().all():
            verificaciones[v.id_factura] = v.verificado

    lista_clientes = []
    monto_total_cobrado = 0.0

    for f in data.get("results", []):
        emision = (f.get("fecha_emision") or "")[:10]
        if not (fecha_str <= emision <= fecha_fin_str):
            continue

        id_factura = f["id_factura"]
        monto = float(f.get("total") or 0)
        estado = f.get("estado", "")
        verificado = verificaciones.get(id_factura, False)
        articulos = f.get("articulos", [])
        cliente = f.get("cliente") or {}

        if verificado or estado == "Pagada":
            monto_total_cobrado += monto

        id_servicio = next(
            (art.get("servicio", {}).get("id_servicio") for art in articulos if art.get("servicio")),
            None
        )
        lista_clientes.append({
            "id_factura": id_factura,
            "fecha_pago": emision,
            "estado": estado,
            "verificado": verificado,
            "monto_individual": monto,
            "metodo_pago": _metodo_pago(f),
            "tipo_cobro": _tipo_cobro(articulos),
            "cliente": {
                "id_servicio": id_servicio,
                "nombre": cliente.get("nombre", "—"),
                "telefono": cliente.get("telefono", "—"),
                "direccion": cliente.get("direccion", "—"),
            },
        })

    lista_clientes.sort(key=lambda x: (0 if x["verificado"] else 1, x["fecha_pago"]))

    return {
        "fecha": fecha_str,
        "fecha_fin": fecha_fin_str,
        "numero_total_pagos": len(lista_clientes),
        "monto_total_cobrado": round(monto_total_cobrado, 2),
        "lista_clientes": lista_clientes,
    }

async def toggle_verificacion(id_factura: int, notas: str | None, db: AsyncSession) -> dict:
    try:
        result = await db.execute(
            select(VerificacionPago)
            .where(VerificacionPago.id_factura == id_factura)
            .with_for_update()
        )
        registro = result.scalar_one_or_none()

        if registro is None:
            registro = VerificacionPago(
                id_factura=id_factura,
                verificado=True,
                fecha_verificacion=datetime.now(),
                notas=notas,
            )
            db.add(registro)
        else:
            registro.verificado = not registro.verificado
            registro.fecha_verificacion = datetime.now() if registro.verificado else None
            registro.notas = notas if notas is not None else registro.notas

        await db.commit()
    except IntegrityError:
        # Dos requests simultáneos en la misma factura nueva: el segundo pierde la carrera.
        # Rollback y devolver el estado que ganó.
        await db.rollback()
        result = await db.execute(
            select(VerificacionPago).where(VerificacionPago.id_factura == id_factura)
        )
        registro = result.scalar_one()

    return {
        "id_factura": id_factura,
        "verificado": registro.verificado,
        "fecha_verificacion": registro.fecha_verificacion.isoformat() if registro.fecha_verificacion else None,
        "notas": registro.notas,
    }

async def get_alertas_cobranza() -> dict:
    today = date.today()

    clientes_data, facturas_data = await asyncio.gather(
        wisphub_client.get_all("/api/clientes/"),
        wisphub_client.get_all("/api/facturas/"),
    )

    # Mapa id_servicio → datos del cliente (solo Activo y Suspendido, no Cancelado)
    clientes: dict[int, dict] = {}
    for c in clientes_data.get("results", []):
        sid = c.get("id_servicio")
        if sid and c.get("estado") != "Cancelado":
            clientes[sid] = c

    # Mapa id_servicio → factura pendiente más próxima.
    # Usamos fecha_vencimiento (= día de corte en WispHub) solo para ordenar y elegir
    # la factura más urgente. Para calcular días vencidos usamos fecha_pago (= día de pago).
    vmap: dict[int, date] = {}           # fecha_vencimiento — para ordenar
    pmap: dict[int, date] = {}           # fecha_pago — referencia real de días vencidos
    vmap_factura: dict[int, int] = {}
    vmap_total: dict[int, float] = {}
    for f in facturas_data.get("results", []):
        if f.get("estado") != "Pendiente de Pago":
            continue
        fv_str = (f.get("fecha_vencimiento") or "")[:10]
        if not fv_str:
            continue
        try:
            fv = date.fromisoformat(fv_str)
        except ValueError:
            continue
        for art in f.get("articulos", []):
            srv_id = (art.get("servicio") or {}).get("id_servicio")
            if srv_id and (srv_id not in vmap or fv < vmap[srv_id]):
                vmap[srv_id] = fv
                vmap_factura[srv_id] = f["id_factura"]
                vmap_total[srv_id] = float(f.get("total") or 0)
                # fecha_pago de la factura es el día que el cliente debe pagar.
                # WispHub devuelve fecha_pago en formato "DD/MM/YYYY HH:MM:SS",
                # distinto al formato ISO de fecha_vencimiento.
                fp_raw = (f.get("fecha_pago") or "").strip()
                fp_date: date | None = None
                if fp_raw:
                    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                        try:
                            fp_date = datetime.strptime(fp_raw[:19], fmt).date()
                            break
                        except ValueError:
                            continue
                pmap[srv_id] = fp_date if fp_date else fv

    grupos: dict[str, list] = {"hoy": [], "dia_1": [], "dia_2": [], "dia_3": [], "mas_de_3": []}

    for srv_id, fv in vmap.items():
        # Usar fecha_pago como referencia; si no existe, caer a fecha_vencimiento
        fecha_pago = pmap.get(srv_id, fv)
        dias = (today - fecha_pago).days
        if dias < 0:
            continue  # aún no vencido

        c = clientes.get(srv_id)
        if not c:
            continue  # servicio sin cliente activo/suspendido, se omite
        item = {
            "id_servicio": srv_id,
            "nombre": c.get("nombre", "—"),
            "telefono": c.get("telefono", "—"),
            "estado": c.get("estado", "—"),
            "fecha_vencimiento": fecha_pago.isoformat(),
            "dias_vencido": dias,
            "id_factura": vmap_factura.get(srv_id),
            "total": vmap_total.get(srv_id),
        }

        if dias == 0:
            grupos["hoy"].append(item)
        elif dias == 1:
            grupos["dia_1"].append(item)
        elif dias == 2:
            grupos["dia_2"].append(item)
        elif dias == 3:
            grupos["dia_3"].append(item)
        else:
            grupos["mas_de_3"].append(item)

    for g in grupos.values():
        g.sort(key=lambda x: x["nombre"])

    return {
        "total": sum(len(v) for v in grupos.values()),
        "hoy":       {"count": len(grupos["hoy"]),       "items": grupos["hoy"]},
        "dia_1":     {"count": len(grupos["dia_1"]),     "items": grupos["dia_1"]},
        "dia_2":     {"count": len(grupos["dia_2"]),     "items": grupos["dia_2"]},
        "dia_3":     {"count": len(grupos["dia_3"]),     "items": grupos["dia_3"]},
        "mas_de_3":  {"count": len(grupos["mas_de_3"]), "items": grupos["mas_de_3"]},
    }

async def get_log_cobranza(fecha: str | None, db: AsyncSession) -> dict:
    fecha_str = fecha or date.today().isoformat()
    inicio = datetime.combine(date.fromisoformat(fecha_str), datetime.min.time())
    fin = datetime.combine(date.fromisoformat(fecha_str), datetime.max.time())
    result = await db.execute(
        select(LogRecordatorio)
        .where(LogRecordatorio.fecha_envio >= inicio, LogRecordatorio.fecha_envio <= fin)
        .order_by(LogRecordatorio.id.desc())
    )
    logs = result.scalars().all()
    return {
        "fecha": fecha_str,
        "total": len(logs),
        "items": [
            {
                "id": r.id,
                "id_cliente": r.id_cliente,
                "id_factura": r.id_factura,
                "dia_tolerancia": r.dia_tolerancia,
                "fecha_envio": r.fecha_envio.isoformat(),
                "exitoso": r.exitoso,
                "respuesta_api": r.respuesta_api,
            }
            for r in logs
        ],
    }

async def registrar_pago(data: dict, db: AsyncSession) -> dict:
    pago = PagoRegistrado(
        id_cliente=data["id_cliente"],
        id_factura=data.get("id_factura"),
        monto=data["monto"],
        metodo_pago=data.get("metodo_pago", "no_especificado"),
        notas=data.get("notas"),
        fecha_pago=datetime.now(),
    )
    db.add(pago)
    await db.commit()
    await db.refresh(pago)
    return {
        "id": pago.id,
        "id_cliente": pago.id_cliente,
        "id_factura": pago.id_factura,
        "monto": pago.monto,
        "fecha_pago": pago.fecha_pago.isoformat(),
        "metodo_pago": pago.metodo_pago,
        "verificado": pago.verificado,
        "notas": pago.notas,
    }


async def get_pagos_dia(fecha: str | None, db: AsyncSession) -> dict:
    fecha_str = fecha or date.today().isoformat()
    inicio = datetime.combine(date.fromisoformat(fecha_str), datetime.min.time())
    fin = datetime.combine(date.fromisoformat(fecha_str), datetime.max.time())
    result = await db.execute(
        select(PagoRegistrado)
        .where(PagoRegistrado.fecha_pago >= inicio, PagoRegistrado.fecha_pago <= fin)
        .order_by(PagoRegistrado.id.desc())
    )
    pagos = result.scalars().all()
    return {
        "fecha": fecha_str,
        "total": len(pagos),
        "items": [
            {
                "id": p.id,
                "id_cliente": p.id_cliente,
                "id_factura": p.id_factura,
                "monto": p.monto,
                "fecha_pago": p.fecha_pago.isoformat(),
                "metodo_pago": p.metodo_pago,
                "verificado": p.verificado,
                "notas": p.notas,
            }
            for p in pagos
        ],
    }

_FORMAS_PAGO_NOMBRES: dict[int, str] = {
    82219: "Efectivo",
    84091: "Tarjeta",
    82222: "Transferencia Bancaria",}

async def pagar_factura_wisphub(id_factura: int, data: dict, db: AsyncSession | None = None) -> dict:
    fecha_pago = data.get("fecha_pago") or datetime.now().strftime("%Y-%m-%d %H:%M")
    payload = {
        "accion": "1",
        "forma_pago": data["forma_pago"],
        "total_cobrado": data["monto"],
        "fecha_pago": fecha_pago,
    }
    try:
        resultado = await wisphub_client.post(
            f"/api/facturas/{id_factura}/registrar-pago/",
            payload=payload,
        )
    except httpx.HTTPStatusError as e:
        try:
            detalle = e.response.json()
            errores = detalle.get("errors") or detalle.get("detail") or [str(e)]
            mensaje = errores[0] if isinstance(errores, list) else str(errores)
        except Exception:
            mensaje = f"WispHub respondió con error {e.response.status_code}"
        raise ValueError(mensaje) from e
    # El pago ya fue registrado en WispHub — el guardado local es secundario
    if db is not None:
        try:
            fecha_pago_real = None
            if data.get("fecha_pago_real"):
                try:
                    fecha_pago_real = datetime.fromisoformat(data["fecha_pago_real"])
                except ValueError:
                    pass
            tipo_pago = (data.get("tipo_pago") or "").strip()
            cuenta = (data.get("cuenta") or "").strip()
            base = tipo_pago or _FORMAS_PAGO_NOMBRES.get(int(data["forma_pago"]), str(data["forma_pago"]))
            metodo_pago_local = f"{base} - {cuenta}" if cuenta else base
            registro = PagoRegistrado(
                id_cliente=str(data.get("id_servicio", "")),
                nombre_cliente=data.get("nombre_cliente"),
                id_factura=id_factura,
                monto=float(data["monto"]),
                metodo_pago=metodo_pago_local,
                fecha_pago=datetime.now(),
                fecha_pago_real=fecha_pago_real,
            )
            db.add(registro)
            await db.commit()
            await db.refresh(registro)
            resultado["pago_id"] = registro.id
        except Exception as e:
            await db.rollback()
            resultado["pago_id"] = None
            resultado["local_error"] = str(e)
    return resultado

async def get_observaciones_batch(entity_type: str, ids: list[int], db: AsyncSession) -> dict:
    """Devuelve {entity_id: notas} para los IDs solicitados."""
    if entity_type == "recoleccion":
        result = await db.execute(
            select(RecoleccionRegistro).where(RecoleccionRegistro.id_servicio.in_(ids))
        )
        return {r.id_servicio: r.notas for r in result.scalars().all() if r.notas}
    if entity_type == "pago":
        result = await db.execute(
            select(PagoRegistrado).where(PagoRegistrado.id.in_(ids))
        )
        return {r.id: r.notas for r in result.scalars().all() if r.notas}
    result = await db.execute(
        select(Observacion).where(
            Observacion.entity_type == entity_type,
            Observacion.entity_id.in_(ids),
        )
    )
    return {r.entity_id: r.notas for r in result.scalars().all() if r.notas}


async def upsert_observacion(entity_type: str, entity_id: int, notas: str, db: AsyncSession) -> dict:
    """Guarda o actualiza la observación. Rutea al modelo correcto según entity_type."""
    if entity_type == "recoleccion":
        result = await db.execute(
            select(RecoleccionRegistro).where(RecoleccionRegistro.id_servicio == entity_id)
        )
        reg = result.scalar_one_or_none()
        if reg:
            reg.notas = notas or None
            reg.updated_at = datetime.now()
        else:
            reg = RecoleccionRegistro(id_servicio=entity_id, estado_equipo="nada_recuperado", notas=notas or None)
            db.add(reg)
    elif entity_type == "pago":
        result = await db.execute(
            select(PagoRegistrado).where(PagoRegistrado.id == entity_id)
        )
        pago = result.scalar_one_or_none()
        if pago:
            pago.notas = notas or None
    else:
        result = await db.execute(
            select(Observacion).where(
                Observacion.entity_type == entity_type,
                Observacion.entity_id == entity_id,
            )
        )
        obs = result.scalar_one_or_none()
        if obs is None:
            obs = Observacion(entity_type=entity_type, entity_id=entity_id, notas=notas or None)
            db.add(obs)
        else:
            obs.notas = notas or None
            obs.updated_at = datetime.now()
    await db.commit()
    return {"entity_type": entity_type, "entity_id": entity_id, "notas": notas}


async def get_formas_pago() -> list:
    data = await wisphub_client.get("/api/formas-de-pago/")
    return [{"id": f["id"], "nombre": f["nombre"]} for f in data.get("results", [])]


async def get_tecnicos() -> list:
    data = await wisphub_client.get("/api/staff/", params={"page_size": 100})
    return [
        {
            "id": s["id"],
            "nombre": s.get("nombre") or s.get("username", ""),
        }
        for s in data.get("results", [])
    ]


async def get_recoleccion(db: AsyncSession) -> dict:
    from sqlalchemy import select
    today = date.today()
    clientes_data, facturas_data = await asyncio.gather(
        wisphub_client.get_all("/api/clientes/"),
        wisphub_client.get_all("/api/facturas/"),
    )
    clientes: dict[int, dict] = {}
    for c in clientes_data.get("results", []):
        sid = c.get("id_servicio")
        if sid:
            clientes[sid] = c
    vmap: dict[int, dict] = {}
    for f in facturas_data.get("results", []):
        if f.get("estado") != "Pendiente de Pago":
            continue
        fv_str = (f.get("fecha_vencimiento") or "")[:10]
        if not fv_str:
            continue
        try:
            fv = date.fromisoformat(fv_str)
        except ValueError:
            continue
        # Parsear fecha_pago igual que en Cobranza
        # Parsear fecha_pago
        fp_raw = (f.get("fecha_pago") or "").strip()
        fp_date = None
        if fp_raw:
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    fp_date = datetime.strptime(fp_raw[:19], fmt).date()
                    break
                except ValueError:
                    continue
        fecha_ref = fp_date if fp_date else fv
        dias = (today - fecha_ref).days
        if dias < 7:
            continue
        for art in f.get("articulos", []):
            srv_id = (art.get("servicio") or {}).get("id_servicio")
            if srv_id and (srv_id not in vmap or fecha_ref < date.fromisoformat(vmap[srv_id]["fecha_vencimiento"])):
                vmap[srv_id] = {
                    "id_factura": f["id_factura"],
                    "fecha_vencimiento": fecha_ref.isoformat(),
                    "dias_vencido": dias,
                    "total": float(f.get("total") or 0),
                }
    items = []
    for srv_id, fdata in vmap.items():
        c = clientes.get(srv_id, {})
        if c.get("estado") == "Cancelado":
            continue
        items.append({
            "id_servicio": srv_id,
            "id_factura": fdata["id_factura"],
            "nombre": c.get("nombre", "—"),
            "direccion": c.get("direccion", "—"),
            "telefono": c.get("telefono", "—"),
            "estado": c.get("estado", "—"),
            "fecha_vencimiento": fdata["fecha_vencimiento"],
            "dias_vencido": fdata["dias_vencido"],
            "total": fdata["total"],
        })

    items.sort(key=lambda x: x["dias_vencido"], reverse=True)

    ids = [item["id_servicio"] for item in items]
    if ids:
        result = await db.execute(
            select(RecoleccionRegistro).where(RecoleccionRegistro.id_servicio.in_(ids))
        )
        registros = {r.id_servicio: r for r in result.scalars().all()}
    else:
        registros = {}

    for item in items:
        reg = registros.get(item["id_servicio"])
        item["estado_equipo"] = reg.estado_equipo if reg else None
        item["notas"] = reg.notas if reg else None
        item["id_tecnico"] = reg.id_tecnico if reg else None
        item["nombre_tecnico"] = reg.nombre_tecnico if reg else None

    return {
        "total": len(items),
        "items": items,
    }

async def guardar_estado_equipo(id_servicio: int, data: dict, db: AsyncSession) -> dict:
    from sqlalchemy import select
    result = await db.execute(
        select(RecoleccionRegistro).where(RecoleccionRegistro.id_servicio == id_servicio)
    )
    registro = result.scalar_one_or_none()

    estado = data.get("estado_equipo", "nada_recuperado")
    notas = data.get("notas")
    id_tecnico = data.get("id_tecnico")
    nombre_tecnico = data.get("nombre_tecnico")

    if registro is None:
        registro = RecoleccionRegistro(
            id_servicio=id_servicio,
            estado_equipo=estado,
            notas=notas,
            id_tecnico=id_tecnico,
            nombre_tecnico=nombre_tecnico,
        )
        db.add(registro)
    else:
        registro.estado_equipo = estado
        registro.notas = notas
        registro.id_tecnico = id_tecnico
        registro.nombre_tecnico = nombre_tecnico
        registro.updated_at = datetime.now()

    await db.commit()
    await db.refresh(registro)
    return {
        "id_servicio": id_servicio,
        "estado_equipo": registro.estado_equipo,
        "notas": registro.notas,
        "id_tecnico": registro.id_tecnico,
        "nombre_tecnico": registro.nombre_tecnico,
        "fecha_registro": registro.fecha_registro.isoformat() if registro.fecha_registro else None,
    }


async def get_historial_pagos(db: AsyncSession, search: str | None = None) -> dict:
    result = await db.execute(select(PagoRegistrado).order_by(PagoRegistrado.fecha_pago.desc()))
    pagos = result.scalars().all()

    if search:
        q = search.lower()
        pagos = [
            p for p in pagos
            if (p.nombre_cliente and q in p.nombre_cliente.lower())
            or q in p.id_cliente
            or (p.id_factura and q in str(p.id_factura))
        ]

    return {
        "total": len(pagos),
        "items": [
            {
                "id": p.id,
                "id_cliente": p.id_cliente,
                "nombre_cliente": p.nombre_cliente,
                "id_factura": p.id_factura,
                "monto": p.monto,
                "metodo_pago": p.metodo_pago,
                "fecha_pago_real": p.fecha_pago_real.isoformat() if p.fecha_pago_real else None,
                "fecha_registro": p.fecha_pago.isoformat(),
                "notas": p.notas,
                "comprobante_url": f"/static/comprobantes/{os.path.basename(p.comprobante_path)}" if p.comprobante_path else None,
            }
            for p in pagos
        ],
    }


_METODO_LABELS: dict[str, str] = {
    "no_especificado": "No especificado",
    "efectivo": "Efectivo",
    "transferencia": "Transferencia",
    "deposito_oxxo": "OXXO",
    "tarjeta": "Tarjeta",
    "transferencia bancaria": "Transferencia Bancaria",
}


async def get_reporte_semanal_data(fecha_inicio: str, fecha_fin: str, db: AsyncSession) -> dict:
    today = date.today()
    facturas_data, clientes_data = await asyncio.gather(
        wisphub_client.get_all("/api/facturas/"),
        wisphub_client.get_all("/api/clientes/"),
    )

    results = facturas_data.get("results", [])

    # Facturas emitidas en el rango
    rango: list[dict] = [
        f for f in results
        if fecha_inicio <= (f.get("fecha_emision") or "")[:10] <= fecha_fin
    ]

    # ── Preparar fechas y pre-cargar PagoRegistrado del rango ─────────────
    from sqlalchemy import or_, and_, func
    date_inicio = date.fromisoformat(fecha_inicio)
    date_fin    = date.fromisoformat(fecha_fin)
    inicio_dt   = datetime.combine(date_inicio, datetime.min.time())
    fin_dt      = datetime.combine(date_fin,    datetime.max.time())

    res_pagos = await db.execute(
        select(PagoRegistrado).where(
            or_(
                and_(
                    PagoRegistrado.fecha_pago_real.isnot(None),
                    func.date(PagoRegistrado.fecha_pago_real).between(date_inicio, date_fin),
                ),
                and_(
                    PagoRegistrado.fecha_pago_real.is_(None),
                    PagoRegistrado.fecha_pago.between(inicio_dt, fin_dt),
                ),
            )
        )
    )
    pagos_locales = res_pagos.scalars().all()
    # Facturas gestionadas localmente: WispHub las excluye de por_dia para que la
    # fecha de registro (fecha_pago) mande, no la fecha de emisión de WispHub.
    pago_local_factura_ids: set[int] = {p.id_factura for p in pagos_locales if p.id_factura}

    # ── Ingresos por día — fuente primaria: WispHub ────────────────────────
    por_dia: dict[str, dict] = {}

    for f in rango:
        emision = (f.get("fecha_emision") or "")[:10]
        if not emision:
            continue
        fid = f.get("id_factura")
        # Si la factura tiene un PagoRegistrado, la fecha correcta vendrá de allí
        if fid and fid in pago_local_factura_ids:
            continue
        monto = float(f.get("total") or 0)
        pagada = f.get("estado") == "Pagada"
        if emision not in por_dia:
            por_dia[emision] = {"fecha": emision, "total_pagado": 0.0, "total_pendiente": 0.0,
                                "count_pagadas": 0, "count_pendientes": 0}
        if pagada:
            por_dia[emision]["total_pagado"] += monto
            por_dia[emision]["count_pagadas"] += 1
        else:
            por_dia[emision]["total_pendiente"] += monto
            por_dia[emision]["count_pendientes"] += 1

    # ── Por método de pago con desglose por cuenta ─────────────────────────
    _metodo_cuenta: dict[str, dict[str, dict]] = {}

    def _add_metodo_cuenta(metodo: str, cuenta: str, monto: float) -> None:
        if metodo not in _metodo_cuenta:
            _metodo_cuenta[metodo] = {}
        if cuenta not in _metodo_cuenta[metodo]:
            _metodo_cuenta[metodo][cuenta] = {"total": 0.0, "count": 0}
        _metodo_cuenta[metodo][cuenta]["total"] += monto
        _metodo_cuenta[metodo][cuenta]["count"] += 1

    # 1) WispHub: facturas Pagadas que NO tienen PagoRegistrado local
    wisphub_factura_ids: set[int] = set()
    for f in rango:
        if f.get("estado") != "Pagada":
            continue
        fid = f.get("id_factura")
        if fid:
            wisphub_factura_ids.add(fid)
        if fid and fid in pago_local_factura_ids:
            continue  # método de pago lo registra PagoRegistrado
        monto = float(f.get("total") or 0)
        fp = f.get("forma_pago") or f.get("metodo_pago") or {}
        if isinstance(fp, dict):
            tipo_id = fp.get("id") or fp.get("Id")
            nombre_cuenta = (fp.get("nombre") or fp.get("Nombre") or "Sin especificar").strip()
            metodo_label = _FORMAS_PAGO_NOMBRES.get(int(tipo_id), "Otro") if tipo_id else "Otro"
        else:
            metodo_label = _METODO_LABELS.get(str(fp).strip().lower(), str(fp).strip() or "Otro")
            nombre_cuenta = "Sin especificar"
        _add_metodo_cuenta(metodo_label, nombre_cuenta, monto)

    # 2) PagoRegistrado local: siempre actualiza por_dia con la fecha de registro
    for p in pagos_locales:
        monto_p = float(p.monto)
        raw_m = (p.metodo_pago or "Otro").strip()
        if " - " in raw_m:
            metodo, cuenta = raw_m.split(" - ", 1)
        else:
            metodo, cuenta = raw_m, "Sin especificar"
        _add_metodo_cuenta(metodo, cuenta, monto_p)

        # Pagos recibidos: usar fecha_pago_real ("Fecha de pago" en UI),
        # o fecha_pago (timestamp de registro) si no se ingresó fecha_pago_real.
        fecha_cobro = (p.fecha_pago_real or p.fecha_pago).date()
        if date_inicio <= fecha_cobro <= date_fin:
            fcobro = fecha_cobro.isoformat()
            if fcobro not in por_dia:
                por_dia[fcobro] = {"fecha": fcobro, "total_pagado": 0.0, "total_pendiente": 0.0,
                                   "count_pagadas": 0, "count_pendientes": 0}
            por_dia[fcobro]["total_pagado"] += monto_p
            por_dia[fcobro]["count_pagadas"] += 1

    por_metodo = []
    for metodo, cuentas in _metodo_cuenta.items():
        cuentas_list = sorted(
            [{"cuenta": k, "total": round(v["total"], 2), "count": v["count"]} for k, v in cuentas.items()],
            key=lambda x: -x["total"],
        )
        por_metodo.append({
            "metodo": metodo,
            "total": round(sum(c["total"] for c in cuentas_list), 2),
            "count": sum(c["count"] for c in cuentas_list),
            "cuentas": cuentas_list,
        })
    por_metodo.sort(key=lambda x: -x["total"])

    # Resumen de clientes — estados del servicio (mutuamente excluyentes)
    clientes = clientes_data.get("results", [])
    total_clientes = len(clientes)

    activos     = sum(1 for c in clientes if c.get("estado") == "Activo")
    suspendidos = sum(1 for c in clientes if c.get("estado") == "Suspendido")
    cancelados  = sum(1 for c in clientes if c.get("estado") == "Cancelado")

    # Indicadores de cartera (informativos — pueden solaparse con estados)
    # Clientes únicos con al menos una factura Pendiente de Pago
    pendientes_pago_ids: set[int] = set()
    en_recoleccion: set[int] = set()
    for f in results:
        if f.get("estado") != "Pendiente de Pago":
            continue
        fv_str = (f.get("fecha_vencimiento") or "")[:10]
        if not fv_str:
            continue
        try:
            fv = date.fromisoformat(fv_str)
        except ValueError:
            continue
        fp_raw = (f.get("fecha_pago") or "").strip()
        fp_date = None
        if fp_raw:
            for fmt_str in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                try:
                    fp_date = datetime.strptime(fp_raw[:19], fmt_str).date()
                    break
                except ValueError:
                    continue
        fecha_ref = fp_date if fp_date else fv
        for art in f.get("articulos", []):
            srv_id = (art.get("servicio") or {}).get("id_servicio")
            if srv_id:
                pendientes_pago_ids.add(srv_id)
                if (today - fecha_ref).days >= 7:
                    en_recoleccion.add(srv_id)

    _estado_cliente: dict[int, str] = {
        c.get("id_servicio"): c.get("estado", "")
        for c in clientes
        if c.get("id_servicio")
    }
    activos_con_deuda     = sum(1 for sid in pendientes_pago_ids if _estado_cliente.get(sid) == "Activo")
    suspendidos_con_deuda = sum(1 for sid in pendientes_pago_ids if _estado_cliente.get(sid) == "Suspendido")

    def pct(n: int, total: int) -> float:
        return round(n / total * 100, 1) if total else 0.0

    total_pagado    = sum(f["total_pagado"]    for f in por_dia.values())
    total_pendiente = sum(f["total_pendiente"] for f in por_dia.values())
    count_pagadas   = sum(f["count_pagadas"]   for f in por_dia.values())

    return {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "total_ingresado": round(total_pagado, 2),
        "total_pendiente": round(total_pendiente, 2),
        "total_facturas_pagadas": count_pagadas,
        "por_dia": sorted(list(por_dia.values()), key=lambda x: x["fecha"]),
        "por_metodo": por_metodo,
        "resumen_clientes": {
            # Sección A: estados del servicio (excluyentes → suman correctamente)
            "total_real": total_clientes,
            "activos": activos,
            "suspendidos": suspendidos,
            "cancelados": cancelados,
            "pct_activos": pct(activos, total_clientes),
            "pct_suspendidos": pct(suspendidos, total_clientes),
            "pct_cancelados": pct(cancelados, total_clientes),
            # Sección B: indicadores de cartera (pueden solaparse con A)
            "pendientes_de_pago": len(pendientes_pago_ids),
            "activos_con_deuda": activos_con_deuda,
            "suspendidos_con_deuda": suspendidos_con_deuda,
            "en_recoleccion": len(en_recoleccion),
            "pct_pendientes_de_pago": pct(len(pendientes_pago_ids), total_clientes),
            "pct_activos_con_deuda": pct(activos_con_deuda, total_clientes),
            "pct_suspendidos_con_deuda": pct(suspendidos_con_deuda, total_clientes),
            "pct_en_recoleccion": pct(len(en_recoleccion), total_clientes),
        },
    }


def _excel_header_style(ws, row: int, cols: int, title: str, fill_color: str = "1E3A8A") -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    end_col = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _excel_col_headers(ws, row: int, headers: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor="E2E8F0")
        c.alignment = Alignment(horizontal="center")


def generar_excel_reporte(data: dict) -> bytes:
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint

    MXN = '#,##0.00'
    GREEN_PALETTE = ["15803D", "22C55E", "4ADE80", "86EFAC", "166534", "BBF7D0"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Semanal"

    for col, w in enumerate([30, 16, 18, 22, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for letter in list("GHIJKLMNOP"):
        ws.column_dimensions[letter].width = 12

    def sec(title: str, color: str, ncols: int = 5) -> int:
        r = ws.max_row + 1
        ws.append([title] + [""] * (ncols - 1))
        ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left")
        return r

    def bold_total(row: int, ncols: int, fgColor: str):
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=fgColor)

    def color_points(series, color_list: list):
        for idx, hex_color in enumerate(color_list):
            pt = DataPoint(idx=idx)
            pt.spPr.solidFill = hex_color
            series.dPt.append(pt)

    # ── Encabezado ──────────────────────────────────────────────────────────
    _excel_header_style(ws, 1, 5,
        f"SIT — Reporte Semanal  ·  {data['fecha_inicio']} al {data['fecha_fin']}")
    ws.append([])

    # ── Ingresos por Día ─────────────────────────────────────────────────────
    sec_dia_start = sec("Ingresos por Día", "1E3A8A")
    _excel_col_headers(ws, ws.max_row + 1,
        ["Fecha", "Fact. cobradas", "Total cobrado", "Fact. pendientes", "Total pendiente"])
    sec_dia_data_start = ws.max_row + 1
    for dia in data["por_dia"]:
        ws.append([dia["fecha"], dia["count_pagadas"], dia["total_pagado"],
                   dia["count_pendientes"], dia["total_pendiente"]])
        for col in (3, 5):
            ws.cell(row=ws.max_row, column=col).number_format = MXN
    sec_dia_data_end = ws.max_row
    tr = ws.max_row + 1
    ws.append(["TOTAL", data["total_facturas_pagadas"], data["total_ingresado"],
               "", data["total_pendiente"]])
    for col in (3, 5):
        ws.cell(row=tr, column=col).number_format = MXN
    bold_total(tr, 5, "DBEAFE")
    ws.append([])

    # ── Por Método / Cuenta ──────────────────────────────────────────────────
    total_count = sum(m["count"] for m in data["por_metodo"])
    total_monto = round(sum(m["total"] for m in data["por_metodo"]), 2)

    sec_met_start = sec("Por Método de Pago", "15803D")
    _excel_col_headers(ws, ws.max_row + 1, ["Método de Pago", "Pagos", "Total cobrado"])
    sec_met_data_start = ws.max_row + 1
    for m in data["por_metodo"]:
        ws.append([m["metodo"], m["count"], m["total"]])
        ws.cell(row=ws.max_row, column=3).number_format = MXN
    sec_met_data_end = ws.max_row
    tr = ws.max_row + 1
    ws.append(["TOTAL", total_count, total_monto])
    ws.cell(row=tr, column=3).number_format = MXN
    bold_total(tr, 3, "DCFCE7")
    ws.append([])

    sec_cta_start = sec("Desglose por Cuenta", "15803D")
    _excel_col_headers(ws, ws.max_row + 1, ["Cuenta", "Pagos", "Total cobrado"])
    sec_cta_data_start = ws.max_row + 1
    for c in sorted([c for m in data["por_metodo"] for c in m["cuentas"]],
                    key=lambda x: -x["total"]):
        ws.append([c["cuenta"], c["count"], c["total"]])
        ws.cell(row=ws.max_row, column=3).number_format = MXN
    sec_cta_data_end = ws.max_row
    tr = ws.max_row + 1
    ws.append(["TOTAL", total_count, total_monto])
    ws.cell(row=tr, column=3).number_format = MXN
    bold_total(tr, 3, "DCFCE7")
    ws.append([])

    # ── Resumen de Clientes ──────────────────────────────────────────────────
    sec_cli_start = sec("Resumen de Clientes", "7C3AED")
    rc = data["resumen_clientes"]
    _excel_col_headers(ws, ws.max_row + 1, ["Indicador", "Clientes", "% del total"])
    sec_cli_data_start = ws.max_row + 1

    for label, qty, pct_val, fill in [
        ("Activos",     rc["activos"],     rc["pct_activos"],     "F0FDF4"),
        ("Suspendidos", rc["suspendidos"], rc["pct_suspendidos"], "FFF7ED"),
        ("Cancelados",  rc["cancelados"],  rc["pct_cancelados"],  "FEF2F2"),
    ]:
        ws.append([label, qty, f"{pct_val}%"])
        for col in range(1, 4):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor=fill)

    total_estado = rc["activos"] + rc["suspendidos"] + rc["cancelados"]
    pct_estado = round(total_estado / rc["total_real"] * 100, 1) if rc["total_real"] else 0
    tr = ws.max_row + 1
    ws.append(["TOTAL", total_estado, f"{pct_estado}%"])
    bold_total(tr, 3, "EDE9FE")
    ws.append([])

    for label, qty, pct_val, fill in [
        ("Activos con deuda",        rc["activos_con_deuda"],     rc["pct_activos_con_deuda"],     "FEF9C3"),
        ("Suspendidos con deuda",    rc["suspendidos_con_deuda"], rc["pct_suspendidos_con_deuda"], "FEF9C3"),
        ("En recolección (7+ días)", rc["en_recoleccion"],        rc["pct_en_recoleccion"],        "EDE9FE"),
    ]:
        ws.append([label, qty, f"{pct_val}%"])
        for col in range(1, 4):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor=fill)

    # ── Gráficas ─────────────────────────────────────────────────────────────
    # Chart 1: BarChart vertical — Ingresos por Día
    c1 = BarChart()
    c1.type = "col"
    c1.title = "Ingresos por Día"
    c1.y_axis.title = "MXN"
    c1.width = 14
    c1.height = 9
    c1.add_data(Reference(ws, min_col=3, min_row=sec_dia_data_start - 1,
                          max_row=sec_dia_data_end), titles_from_data=True)
    c1.set_categories(Reference(ws, min_col=1, min_row=sec_dia_data_start,
                                max_row=sec_dia_data_end))
    n_dias = sec_dia_data_end - sec_dia_data_start + 1
    color_points(c1.series[0], ["1E3A8A"] * n_dias)
    ws.add_chart(c1, f"G{sec_dia_start}")

    # Chart 2: DoughnutChart — Distribución por Método
    c2 = DoughnutChart()
    c2.title = "Distribución por Método"
    c2.holeSize = 50
    c2.width = 12
    c2.height = 9
    c2.add_data(Reference(ws, min_col=3, min_row=sec_met_data_start,
                          max_row=sec_met_data_end))
    c2.set_categories(Reference(ws, min_col=1, min_row=sec_met_data_start,
                                max_row=sec_met_data_end))
    n_met = sec_met_data_end - sec_met_data_start + 1
    color_points(c2.series[0], [GREEN_PALETTE[i % len(GREEN_PALETTE)] for i in range(n_met)])
    ws.add_chart(c2, f"G{sec_met_start}")

    # Chart 3: BarChart horizontal — Top Cuentas
    c3 = BarChart()
    c3.type = "bar"
    c3.title = "Top Cuentas"
    c3.x_axis.title = "MXN"
    n_cta = sec_cta_data_end - sec_cta_data_start + 1
    c3.width = 14
    c3.height = max(7.0, 1.2 * n_cta)
    c3.add_data(Reference(ws, min_col=3, min_row=sec_cta_data_start - 1,
                          max_row=sec_cta_data_end), titles_from_data=True)
    c3.set_categories(Reference(ws, min_col=1, min_row=sec_cta_data_start,
                                max_row=sec_cta_data_end))
    color_points(c3.series[0], ["15803D"] * n_cta)
    ws.add_chart(c3, f"G{sec_cta_start}")

    # Chart 4: PieChart — Estado de Clientes
    c4 = PieChart()
    c4.title = "Estado de Clientes"
    c4.width = 12
    c4.height = 9
    c4.add_data(Reference(ws, min_col=2, min_row=sec_cli_data_start,
                          max_row=sec_cli_data_start + 2))
    c4.set_categories(Reference(ws, min_col=1, min_row=sec_cli_data_start,
                                max_row=sec_cli_data_start + 2))
    color_points(c4.series[0], ["15803D", "D97706", "DC2626"])
    ws.add_chart(c4, f"G{sec_cli_start}")

    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_pdf_reporte(data: dict) -> bytes:
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Flowable, KeepTogether
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics import renderPDF

    PRIMARY   = colors.HexColor("#1E3A8A")
    GREEN     = colors.HexColor("#15803D")
    PURPLE    = colors.HexColor("#7C3AED")
    LIGHT     = colors.HexColor("#F8FAFC")
    GRID_CLR  = colors.HexColor("#CBD5E1")
    ALT_ROW   = colors.HexColor("#F1F5F9")
    NARANJA   = colors.HexColor("#D97706")
    ROJO      = colors.HexColor("#DC2626")
    GREEN_SHADES = [
        colors.HexColor("#15803D"), colors.HexColor("#22C55E"),
        colors.HexColor("#4ADE80"), colors.HexColor("#86EFAC"),
        colors.HexColor("#166534"), colors.HexColor("#BBF7D0"),
    ]

    _PAGE_W = 7.0 * inch  # ancho disponible (8.5" - 2 * 0.75" márgenes)

    class _ChartFlowable(Flowable):
        def __init__(self, drawing: Drawing):
            super().__init__()
            self.drawing = drawing
            self.width = _PAGE_W
            self.height = drawing.height
        def draw(self):
            x = (_PAGE_W - self.drawing.width) / 2
            renderPDF.draw(self.drawing, self.canv, x, 0)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("sit_title", parent=styles["Title"],
                                 textColor=PRIMARY, fontSize=20, spaceAfter=2)
    sub_style = ParagraphStyle("sit_sub", parent=styles["Normal"],
                               textColor=colors.HexColor("#64748B"), fontSize=11, spaceAfter=8)
    section_style = ParagraphStyle("sit_section", parent=styles["Heading2"],
                                   textColor=PRIMARY, fontSize=13, spaceBefore=14, spaceAfter=6)

    def make_table(header: list, rows: list, col_widths: list, accent: object) -> Table:
        all_rows = [header] + rows
        t = Table(all_rows, colWidths=col_widths)
        row_count = len(all_rows)
        ts = TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  accent),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING",    (0, 0), (-1, 0), 8),
            ("ALIGN",       (1, 0), (-1, -1), "CENTER"),
            ("FONTSIZE",    (0, 1), (-1, -1), 9),
            ("TOPPADDING",  (0, 1), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ("GRID",        (0, 0), (-1, -1), 0.4, GRID_CLR),
        ])
        for i in range(1, row_count):
            bg = ALT_ROW if i % 2 == 0 else colors.white
            ts.add("BACKGROUND", (0, i), (-1, i), bg)
        t.setStyle(ts)
        return t

    total_local_count = sum(m["count"] for m in data["por_metodo"])
    total_local_monto = sum(m["total"] for m in data["por_metodo"])
    todas_cuentas_pdf = sorted(
        [c for m in data["por_metodo"] for c in m["cuentas"]],
        key=lambda x: -x["total"],
    )
    rc = data["resumen_clientes"]
    total_estado = rc["activos"] + rc["suspendidos"] + rc["cancelados"]
    pct_estado = round(total_estado / rc["total_real"] * 100, 1) if rc["total_real"] else 0

    elems = []

    # ── Encabezado ─────────────────────────────────────────────────────────
    elems.append(Paragraph("SIT — Panel de Control", title_style))
    elems.append(Paragraph(
        f"Reporte de Ingresos: {data['fecha_inicio']} al {data['fecha_fin']}", sub_style))
    elems.append(HRFlowable(width="100%", thickness=1.5, color=PRIMARY, spaceAfter=10))

    kpi_t = Table(
        [["Total ingresado", "Total de Pagos", "Total pendiente"],
         [f"${data['total_ingresado']:,.2f}", str(data["total_facturas_pagadas"]),
          f"${data['total_pendiente']:,.2f}"]],
        colWidths=[2.2 * inch, 2 * inch, 2.2 * inch],
    )
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), LIGHT),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.HexColor("#475569")),
        ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 1), (-1, 1), 14),
        ("TEXTCOLOR",   (0, 1), (0, 1),  PRIMARY),
        ("TEXTCOLOR",   (2, 1), (2, 1),  colors.HexColor("#D97706")),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING",  (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX",         (0, 0), (-1, -1), 1, GRID_CLR),
        ("INNERGRID",   (0, 0), (-1, -1), 0.5, GRID_CLR),
        ("ROUNDEDCORNERS", [6]),
    ]))
    elems.append(kpi_t)
    elems.append(Spacer(1, 0.25 * inch))

    # ── Tablas ─────────────────────────────────────────────────────────────
    elems.append(Paragraph("Ingresos por Día", section_style))
    dia_rows = [
        [d["fecha"], str(d["count_pagadas"]), f"${d['total_pagado']:,.2f}",
         str(d["count_pendientes"]), f"${d['total_pendiente']:,.2f}"]
        for d in data["por_dia"]
    ]
    dia_rows.append(["TOTAL", str(data["total_facturas_pagadas"]),
                     f"${data['total_ingresado']:,.2f}", "", f"${data['total_pendiente']:,.2f}"])
    dia_t = make_table(
        ["Fecha", "Pagos recibidos", "Total cobrado", "Pagos pendientes", "Total pendiente"],
        dia_rows, [1.4 * inch, 1.1 * inch, 1.5 * inch, 1.1 * inch, 1.5 * inch], PRIMARY,
    )
    dia_t.setStyle(TableStyle([
        ("FONTNAME",   (0, len(dia_rows)), (-1, len(dia_rows)), "Helvetica-Bold"),
        ("BACKGROUND", (0, len(dia_rows)), (-1, len(dia_rows)), colors.HexColor("#DBEAFE")),
    ]))
    elems.append(dia_t)
    elems.append(Spacer(1, 0.2 * inch))

    elems.append(Paragraph("Ingresos por Método de Pago", section_style))
    met1_rows = [[m["metodo"], str(m["count"]), f"${m['total']:,.2f}"] for m in data["por_metodo"]]
    met1_rows.append(["TOTAL", str(total_local_count), f"${total_local_monto:,.2f}"])
    met1_t = make_table(
        ["Método de Pago", "Pagos", "Total cobrado"],
        met1_rows, [2.5 * inch, 1.0 * inch, 2.7 * inch], GREEN,
    )
    met1_t.setStyle(TableStyle([
        ("FONTNAME",   (0, len(met1_rows)), (-1, len(met1_rows)), "Helvetica-Bold"),
        ("BACKGROUND", (0, len(met1_rows)), (-1, len(met1_rows)), colors.HexColor("#DCFCE7")),
    ]))
    elems.append(met1_t)
    elems.append(Spacer(1, 0.15 * inch))

    elems.append(Paragraph("Desglose por Cuenta", section_style))
    met2_rows = [[c["cuenta"], str(c["count"]), f"${c['total']:,.2f}"] for c in todas_cuentas_pdf]
    met2_rows.append(["TOTAL", str(total_local_count), f"${total_local_monto:,.2f}"])
    met2_t = make_table(
        ["Cuenta", "Pagos", "Total cobrado"],
        met2_rows, [2.8 * inch, 0.9 * inch, 1.6 * inch], GREEN,
    )
    met2_t.setStyle(TableStyle([
        ("FONTNAME",   (0, len(met2_rows)), (-1, len(met2_rows)), "Helvetica-Bold"),
        ("BACKGROUND", (0, len(met2_rows)), (-1, len(met2_rows)), colors.HexColor("#DCFCE7")),
    ]))
    elems.append(met2_t)
    elems.append(Spacer(1, 0.2 * inch))

    elems.append(Paragraph("Resumen de Clientes", section_style))
    cli_rows = [
        ["Activos",     str(rc["activos"]),     f"{rc['pct_activos']}%"],
        ["Suspendidos", str(rc["suspendidos"]),  f"{rc['pct_suspendidos']}%"],
        ["Cancelados",  str(rc["cancelados"]),   f"{rc['pct_cancelados']}%"],
        ["TOTAL",       str(total_estado),        f"{pct_estado}%"],
    ]
    cli_t = make_table(
        ["Estado del Servicio", "Clientes", "% del total"],
        cli_rows, [2.8 * inch, 1.3 * inch, 2.1 * inch], PURPLE,
    )
    cli_t.setStyle(TableStyle([
        ("FONTNAME",   (0, len(cli_rows)), (-1, len(cli_rows)), "Helvetica-Bold"),
        ("BACKGROUND", (0, len(cli_rows)), (-1, len(cli_rows)), colors.HexColor("#EDE9FE")),
    ]))
    elems.append(cli_t)
    elems.append(Spacer(1, 0.15 * inch))

    AMBER = colors.HexColor("#B45309")
    ind_rows = [
        ["Activos con deuda",        str(rc["activos_con_deuda"]),     f"{rc['pct_activos_con_deuda']}%"],
        ["Suspendidos con deuda",    str(rc["suspendidos_con_deuda"]), f"{rc['pct_suspendidos_con_deuda']}%"],
        ["En recolección (7+ días)", str(rc["en_recoleccion"]),        f"{rc['pct_en_recoleccion']}%"],
    ]
    ind_t = make_table(
        ["Indicadores de Cartera", "Clientes", "% del total"],
        ind_rows, [2.8 * inch, 1.3 * inch, 2.1 * inch], AMBER,
    )
    elems.append(ind_t)
    elems.append(Spacer(1, 0.3 * inch))

    # ── Gráficas ───────────────────────────────────────────────────────────
    elems.append(HRFlowable(width="100%", thickness=1, color=GRID_CLR, spaceAfter=6))
    elems.append(Paragraph("Gráficas", section_style))

    # Chart 1: Ingresos por Día
    if data["por_dia"]:
        _pw = 7.0 * inch
        d = Drawing(_pw, 2.8 * inch)
        bc = VerticalBarChart()
        bc.x, bc.y = 45, 20
        bc.width, bc.height = _pw - 60, 2.2 * inch
        bc.data = [[dia["total_pagado"] for dia in data["por_dia"]]]
        bc.categoryAxis.categoryNames = [dia["fecha"] for dia in data["por_dia"]]
        bc.categoryAxis.labels.angle = 30 if len(data["por_dia"]) > 4 else 0
        bc.categoryAxis.labels.fontSize = 7
        bc.valueAxis.labelTextFormat = lambda v: f"${v:,.0f}"
        bc.valueAxis.labels.fontSize = 7
        bc.bars[0].fillColor = PRIMARY
        bc.bars[0].strokeColor = PRIMARY
        bc.groupSpacing = 5
        d.add(bc)
        elems.append(KeepTogether([
            Paragraph("Ingresos por Día", section_style),
            _ChartFlowable(d),
            Spacer(1, 0.2 * inch),
        ]))

    # Chart 2: Distribución por Método
    if data["por_metodo"] and total_local_monto:
        _pie_sz = 3.8 * inch
        d = Drawing(_pie_sz, _pie_sz)
        pie = Pie()
        pie.x, pie.y = 50, 30
        pie.width = pie.height = _pie_sz - 80
        pie.data = [m["total"] for m in data["por_metodo"]]
        pie.labels = [
            f"{m['metodo']}\n${m['total']:,.0f} ({round(m['total']/total_local_monto*100,1)}%)"
            for m in data["por_metodo"]
        ]
        pie.sideLabels = True
        pie.simpleLabels = False
        pie.sideLabelsOffset = 0.08
        for i, shade in enumerate(GREEN_SHADES[:len(data["por_metodo"])]):
            pie.slices[i].fillColor = shade
        pie.slices.strokeColor = colors.white
        pie.slices.strokeWidth = 0.5
        d.add(pie)
        elems.append(KeepTogether([
            Paragraph("Distribución por Método de Pago", section_style),
            _ChartFlowable(d),
            Spacer(1, 0.2 * inch),
        ]))

    # Chart 3: Top Cuentas
    top_cuentas = todas_cuentas_pdf[:8]
    if top_cuentas:
        _ch = max(2.0 * inch, 0.35 * inch * len(top_cuentas))
        d = Drawing(7.0 * inch, _ch + 0.5 * inch)
        hbc = HorizontalBarChart()
        hbc.x, hbc.y = 110, 20
        hbc.width = 7.0 * inch - 130
        hbc.height = _ch
        hbc.data = [[c["total"] for c in top_cuentas]]
        hbc.categoryAxis.categoryNames = [c["cuenta"] for c in top_cuentas]
        hbc.categoryAxis.labels.fontSize = 7
        hbc.categoryAxis.labels.dx = -5
        hbc.valueAxis.labelTextFormat = lambda v: f"${v:,.0f}"
        hbc.valueAxis.labels.fontSize = 7
        hbc.bars[0].fillColor = GREEN
        hbc.bars[0].strokeColor = GREEN
        hbc.reversePlotOrder = True
        d.add(hbc)
        elems.append(KeepTogether([
            Paragraph("Top Cuentas por Monto", section_style),
            _ChartFlowable(d),
            Spacer(1, 0.2 * inch),
        ]))

    # Chart 4: Estado de Clientes
    rc_data = [
        (rc["activos"],     "Activos",     rc["pct_activos"],     GREEN),
        (rc["suspendidos"], "Suspendidos", rc["pct_suspendidos"], NARANJA),
        (rc["cancelados"],  "Cancelados",  rc["pct_cancelados"],  ROJO),
    ]
    visible = [(v, l, pct, c) for v, l, pct, c in rc_data if v > 0]
    if visible:
        _pie_sz = 3.8 * inch
        d = Drawing(_pie_sz, _pie_sz)
        pie = Pie()
        pie.x, pie.y = 50, 30
        pie.width = pie.height = _pie_sz - 80
        pie.data   = [v for v, _, _, _ in visible]
        pie.labels = [f"{l}\n{v} ({pct}%)" for v, l, pct, _ in visible]
        pie.sideLabels = True
        pie.simpleLabels = False
        pie.sideLabelsOffset = 0.08
        for i, (_, _, _, color) in enumerate(visible):
            pie.slices[i].fillColor = color
        pie.slices.strokeColor = colors.white
        pie.slices.strokeWidth = 0.5
        d.add(pie)
        elems.append(KeepTogether([
            Paragraph("Estado de Clientes", section_style),
            _ChartFlowable(d),
            Spacer(1, 0.2 * inch),
        ]))

    # ── Pie ────────────────────────────────────────────────────────────────
    elems.append(HRFlowable(width="100%", thickness=0.5, color=GRID_CLR))
    elems.append(Paragraph(
        f"Generado el {date.today().isoformat()}  ·  SIT v1.2 — Panel de Control",
        ParagraphStyle("footer", parent=styles["Normal"],
                       textColor=colors.HexColor("#94A3B8"), fontSize=8, spaceAfter=0),
    ))

    doc.build(elems)
    return buf.getvalue()


async def upload_comprobante(pago_id: int, filename: str, content: bytes, db: AsyncSession) -> dict:
    ext = os.path.splitext(filename)[1].lower()
    safe_name = f"{pago_id}_{uuid.uuid4().hex}{ext}"
    path = os.path.join("data", "comprobantes", safe_name)

    with open(path, "wb") as f:
        f.write(content)

    result = await db.execute(select(PagoRegistrado).where(PagoRegistrado.id == pago_id))
    registro = result.scalar_one_or_none()
    if registro:
        registro.comprobante_path = path
        await db.commit()

    return {"comprobante_url": f"/static/comprobantes/{safe_name}"}